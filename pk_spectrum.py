import numpy as np
from scipy.optimize import nnls
from openpyxl import load_workbook

Kw = 1e-14
LOG10 = np.log(10)
TOLERANCE = 1e-6


class pKSpectrum:
    def __init__(self, source_file):
        self.source_file = source_file
        self.sample_name = None
        self.comment = None
        self.date = None
        self.time = None
        self.sample_volume = None
        self.alkaline_concentration = None
        self.alkaline_volumes = []
        self.ph_values = []
        self.alpha_values = []
        self.valid_points = None
        self.acid_peaks = []

        self._load_data()

    def _load_data(self):

        # Load workbook
        wb = load_workbook(self.source_file)
        ws = wb.active

        # Get sample information
        self.sample_name = ws['A1'].value
        self.comment = ws['A2'].value
        self.timestamp = ws['A3'].value
        self.sample_volume = ws['A4'].value
        self.alkaline_concentration = ws['A5'].value

        # Get titration data
        shift = 0
        while True:
            volume = ws[f'A{6 + shift}'].value
            ph = ws[f'B{6 + shift}'].value

            if self._check_number(volume) and self._check_number(ph):
                self.alkaline_volumes.append(volume)
                self.ph_values.append(ph)
                shift += 1
            else:
                break

        # Arrange titration data
        swapped = False
        while True:
            for i in range(len(self.alkaline_volumes)-1):
                if self.alkaline_volumes[i] > self.alkaline_volumes[i+1]:
                    swapped = True
                    self.alkaline_volumes[i], self.alkaline_volumes[i+1] = \
                        self.alkaline_volumes[i+1], self.alkaline_volumes[i]
                    self.ph_values[i], self.ph_values[i+1] = self.ph_values[i+1], self.ph_values[i]
            if not swapped:
                break

        # Check data validity
        for i in range(len(self.alkaline_volumes)):
            h = pow(10, -self.ph_values[i])
            t = ((h - Kw / h) / self.sample_volume) * (self.alkaline_volumes[i] + self.sample_volume) + \
                self.alkaline_concentration * self.alkaline_volumes[i] / self.sample_volume
            if t >= 0:
                self.alpha_values.append(t)
                self.valid_points = i + 1
            else:
                break

    def make_calculation(self, pk_start, pk_end, d_pk):

        # Calculate constant step
        pk_step = round((pk_end - pk_start) / d_pk) + 1

        # Fill right part
        right = np.zeros((self.valid_points, pk_step))
        for i in range(self.valid_points):
            for j in range(pk_step):
                right[i, j] = d_pk / (1 + np.exp(LOG10 * (pk_start + d_pk * j - self.ph_values[i])))

        # Solve equation
        constants, residual = nnls(right, np.array(self.alpha_values))

        # Normalization
        constants *= d_pk

        # Truncate border artefacts
        if constants[0] > TOLERANCE > constants[1]:
            constants[0] = 0
        if constants[-1] > TOLERANCE > constants[-2]:
            constants[-1] = 0

        sum_constants = constants.sum()
        max_constant = constants.max(initial=0)
        threshold = max_constant / 100
        constants_relative = constants / sum_constants

        # Peak calculation sequence
        i = 0
        while i < pk_step:
            if constants[i] > threshold:
                self.acid_peaks.append({'point_count': 0, 'concentration': 0, 'first_point': i})
                while i < pk_step and constants[i] > threshold:
                    self.acid_peaks[-1]['point_count'] += 1
                    self.acid_peaks[-1]['concentration'] += constants[i]
                    i += 1
            else:
                i += 1

        # Peaks exact position and height calculation
        if len(self.acid_peaks) > 0:
            for i in range(len(self.acid_peaks)):
                t1 = 0
                t2 = 0
                peak = self.acid_peaks[i]
                for j in range(peak['point_count']):
                    t1 += constants_relative[peak['first_point'] + j] * \
                        (pk_start + d_pk * (peak['first_point'] + j))
                    t2 += constants_relative[peak['first_point'] + j]
                peak['mean'] = t1 / t2
            for i in range(len(self.acid_peaks)):
                peak = self.acid_peaks[i]
                if peak['point_count'] > 0:
                    t1 = 0
                    t2 = 0
                    for j in range(peak['point_count']):
                        t1 += constants_relative[peak['first_point'] + j] * \
                              (pk_start + d_pk * (peak['first_point'] + j) - peak['mean']) ** 2
                        t2 += constants_relative[peak['first_point'] + j]
                    peak['interval'] = 1.96 * np.sqrt(t1 / t2) / np.sqrt(peak['point_count'])
                else:
                    peak['interval'] = 0.

        # Calculate error
        error = np.sqrt(residual) / np.sqrt(pk_step - 1)

        return self.acid_peaks, error

    @staticmethod
    def _check_number(a):
        return type(a) == int or type(a) == float
